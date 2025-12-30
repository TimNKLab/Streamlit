"""Stock card generation logic module."""

from __future__ import annotations

import calendar
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple, Union
import zipfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

ExcelSource = Union[str, Path, bytes, bytearray, BytesIO]


class StockCardGenerator:
    """Generates stock cards grouped by parent brand with print settings."""

    def __init__(self) -> None:
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def apply_border_to_range(
        self, ws, start_row: int, end_row: int, start_col: int, end_col: int
    ) -> None:
        """Apply border to all cells in a range."""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = self.thin_border

    def group_by_parent_brand(self, df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """Group data by Parent Brand with Hebe and Paragon exceptions."""
        groups: Dict[str, List[int]] = {}
        split_by_brand_parents = {"Paragon", "Hebe"}

        df_with_brands = df.copy()

        for idx, row in df_with_brands.iterrows():
            parent_brand = (
                row["Parent Brand"] if pd.notna(row["Parent Brand"]) else row["Brand Name"]
            )
            brand = row["Brand Name"] if pd.notna(row["Brand Name"]) else "Unknown"

            if parent_brand in split_by_brand_parents:
                group_key = f"{parent_brand}_{brand}"
            else:
                group_key = parent_brand

            groups.setdefault(group_key, []).append(idx)

        grouped_dfs: Dict[str, pd.DataFrame] = {}
        for key, indices in groups.items():
            grouped_dfs[key] = df_with_brands.loc[indices].reset_index(drop=True)

        return grouped_dfs

    def get_month_dates(self, year: int, month: int) -> List[int]:
        """Get all dates for a given month."""
        num_days = calendar.monthrange(year, month)[1]
        return list(range(1, num_days + 1))

    def create_stock_card_sheet(
        self,
        ws,
        df: pd.DataFrame,
        year: int,
        month: int,
        start_date: int,
        end_date: int,
    ) -> None:
        """Create stock card sheet with proper formatting and print settings."""
        dates = list(range(start_date, end_date + 1))

        df_original = df.copy()
        start_row = 4

        num_dates = len(dates)
        total_cols = 3 + (num_dates * 2)
        last_col_num = total_cols
        last_col_letter = get_column_letter(last_col_num)

        ws.merge_cells("A1:A3")
        cell = ws["A1"]
        cell.value = "Barcode"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        self.apply_border_to_range(ws, 1, 3, 1, 1)

        ws.merge_cells("B1:B3")
        cell = ws["B1"]
        cell.value = "Name"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        self.apply_border_to_range(ws, 1, 3, 2, 2)

        ws.merge_cells("C1:C2")
        cell = ws["C1"]
        cell.value = "Stok Display\nTanggal 1"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.apply_border_to_range(ws, 1, 2, 3, 3)

        ws["C3"] = ""
        ws["C3"].border = self.thin_border

        current_col = 4
        for date in dates:
            col1 = get_column_letter(current_col)
            col2 = get_column_letter(current_col + 1)

            ws.merge_cells(f"{col1}1:{col2}2")

            cell = ws[f"{col1}1"]
            cell.value = f"{date} "
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            self.apply_border_to_range(ws, 1, 2, current_col, current_col + 1)

            ws[f"{col1}3"] = "MSK"
            ws[f"{col1}3"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{col1}3"].font = Font(bold=True, size=9)
            ws[f"{col1}3"].border = self.thin_border

            ws[f"{col2}3"] = "KLR"
            ws[f"{col2}3"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{col2}3"].font = Font(bold=True, size=9)
            ws[f"{col2}3"].border = self.thin_border

            current_col += 2

        current_row = start_row
        for _, row in df_original.iterrows():
            barcode_val = str(row["Barcode"]) if pd.notna(row["Barcode"]) else ""
            ws[f"A{current_row}"] = barcode_val
            ws[f"A{current_row}"].number_format = "@"
            ws[f"A{current_row}"].border = self.thin_border
            ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center")

            ws[f"B{current_row}"] = row["Name"]
            ws[f"B{current_row}"].border = self.thin_border
            ws[f"B{current_row}"].alignment = Alignment(horizontal="left", vertical="center")

            ws[f"C{current_row}"] = (
                row["Stok Display Target"] if pd.notna(row["Stok Display Target"]) else ""
            )
            ws[f"C{current_row}"].border = self.thin_border
            ws[f"C{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            current_col = 4
            for _ in dates:
                col_msk = get_column_letter(current_col)
                col_klr = get_column_letter(current_col + 1)

                cell_msk = ws[f"{col_msk}{current_row}"]
                cell_msk.value = ""
                cell_msk.border = self.thin_border
                cell_msk.alignment = Alignment(horizontal="center", vertical="center")

                cell_klr = ws[f"{col_klr}{current_row}"]
                cell_klr.value = ""
                cell_klr.border = self.thin_border
                cell_klr.alignment = Alignment(horizontal="center", vertical="center")

                current_col += 2

            current_row += 1

        last_data_row = current_row - 1

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 10

        for i in range(4, total_cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 6

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 20

        print_area = f"A1:{last_col_letter}{last_data_row}"
        ws.print_area = print_area
        ws.print_title_cols = "A:C"
        ws.print_title_rows = "1:3"

        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        ws.page_margins = PageMargins(
            left=0.2,
            right=0.2,
            top=0.3,
            bottom=0.3,
            header=0.1,
            footer=0.1,
        )

        ws.print_options.horizontalCentered = True

    def process_stock_data(
        self, excel_file: ExcelSource, year: Optional[int] = None, month: Optional[int] = None
    ) -> Tuple[Dict[str, bytes], bytes, Dict[str, Union[int, str, List[str]]]]:
        """Main processing function for stock cards."""
        if excel_file is None:
            raise ValueError("An Excel file must be provided.")

        try:
            df = self._load_dataframe(excel_file)
        except Exception as exc:
            raise ValueError(f"Failed to read Excel file: {exc}") from exc

        column_mapping = {
            "Product/Barcode": "Barcode",
            "Product/Name": "Name",
            "Product/Brand/Parent Brand": "Parent Brand",
            "Product/Brand/Brand Name": "Brand Name",
            "Quantity": "Stok Display Target",
        }

        df = df.rename(columns=column_mapping)

        required_cols = ["Barcode", "Name", "Parent Brand", "Brand Name", "Stok Display Target"]
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            raise ValueError(f"Missing columns: {', '.join(missing)}")

        now = datetime.now()
        year = year or now.year
        month = month or now.month

        grouped_dfs = self.group_by_parent_brand(df)
        num_days = calendar.monthrange(year, month)[1]

        date_ranges: List[Tuple[int, int]] = []
        if num_days >= 10:
            date_ranges.append((1, 10))
        if num_days >= 20:
            date_ranges.append((11, 20))
        if num_days >= 21:
            date_ranges.append((21, num_days))

        if not date_ranges:
            date_ranges.append((1, num_days))

        workbooks_dict: Dict[str, bytes] = {}

        for group_key, df_group in grouped_dfs.items():
            df_group = df_group[["Barcode", "Name", "Stok Display Target"]].copy()

            wb = Workbook()
            wb.remove(wb.active)

            for start_date, end_date in date_ranges:
                sheet_name = f"{start_date}-{end_date} {calendar.month_abbr[month]}"
                ws = wb.create_sheet(title=sheet_name)
                self.create_stock_card_sheet(ws, df_group, year, month, start_date, end_date)

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"StockCard_{group_key}_{month:02d}{year}.xlsx"
            workbooks_dict[filename] = output.getvalue()

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for filename, file_bytes in workbooks_dict.items():
                zip_file.writestr(filename, file_bytes)

        zip_buffer.seek(0)

        summary = {
            "total_products": len(df),
            "groups_count": len(grouped_dfs),
            "groups": list(grouped_dfs.keys()),
            "year": year,
            "month": month,
            "month_name": calendar.month_name[month],
            "sheets_per_workbook": len(date_ranges),
        }

        return workbooks_dict, zip_buffer.getvalue(), summary

    def _load_dataframe(self, excel_file: ExcelSource) -> pd.DataFrame:
        """Load dataframe from any supported excel source."""
        if isinstance(excel_file, (str, Path)):
            return pd.read_excel(excel_file, dtype={"Product/Barcode": str})

        if isinstance(excel_file, (bytes, bytearray)):
            buffer = BytesIO(excel_file)
            return pd.read_excel(buffer, dtype={"Product/Barcode": str})

        if isinstance(excel_file, BytesIO):
            excel_file.seek(0)
            return pd.read_excel(excel_file, dtype={"Product/Barcode": str})

        if hasattr(excel_file, "read"):
            if hasattr(excel_file, "seek"):
                excel_file.seek(0)
            buffer = BytesIO(excel_file.read())
            buffer.seek(0)
            return pd.read_excel(buffer, dtype={"Product/Barcode": str})

        raise TypeError("Unsupported excel_file type")


__all__ = ["StockCardGenerator"]
