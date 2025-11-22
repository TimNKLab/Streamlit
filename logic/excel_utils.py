"""Excel utilities and formatting functions"""

import pandas as pd
import io
import zipfile
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def sanitize_filename(name):
    """Sanitize filename by removing invalid characters"""
    invalid_chars = r'[<>:"/\\|?*]'
    sanitized = re.sub(invalid_chars, '_', str(name))
    sanitized = sanitized.strip(' .')
    return sanitized if sanitized else 'Unknown'

def create_pivot_by_barcode(df_group):
    """Create pivot table with barcode as index, dates as columns, showing Quantity and Tax Incl."""
    df = df_group.copy()
    
    df['Product/Barcode'] = df['Product/Barcode'].astype(str)
    
    if not pd.api.types.is_datetime64_any_dtype(df['Order Date']):
        df['Order Date'] = pd.to_datetime(df['Order Date'])
    
    df['Order Date Day'] = pd.to_datetime(df['Order Date']).dt.normalize()
    
    pivot_qty = pd.pivot_table(
        df,
        index=['Product/Barcode', 'Product'],
        columns='Order Date Day',
        values='Quantity',
        aggfunc='sum',
        fill_value=0
    )
    
    pivot_revenue = pd.pivot_table(
        df,
        index=['Product/Barcode', 'Product'],
        columns='Order Date Day',
        values='Tax Incl.',
        aggfunc='sum',
        fill_value=0
    )
    
    def format_col_name(prefix, col_val):
        if pd.notna(col_val) and isinstance(col_val, pd.Timestamp):
            return f"{prefix}_{col_val.strftime('%Y-%m-%d')}"
        else:
            return f"{prefix}_{str(col_val)}"
    
    pivot_qty.columns = [format_col_name("Jumlah", col) for col in pivot_qty.columns]
    pivot_revenue.columns = [format_col_name("Total", col) for col in pivot_revenue.columns]
    
    pivot = pd.merge(
        pivot_qty.reset_index(),
        pivot_revenue.reset_index(),
        on=['Product/Barcode', 'Product'],
        how='outer'
    )
    
    pivot = pivot.rename(columns={'Product/Barcode': 'Barcode', 'Product': 'Produk'})
    
    base_cols = ['Barcode', 'Produk']
    qty_cols = [col for col in pivot.columns if col.startswith('Jumlah_')]
    revenue_cols = [col for col in pivot.columns if col.startswith('Total_')]
    
    def extract_date(col_name):
        try:
            date_str = col_name.split('_', 1)[1]
            return pd.to_datetime(date_str)
        except:
            return pd.Timestamp.min
    
    qty_cols_sorted = sorted(qty_cols, key=extract_date)
    revenue_cols_sorted = sorted(revenue_cols, key=extract_date)
    
    pivot = pivot[base_cols + qty_cols_sorted + revenue_cols_sorted]
    
    return pivot

def create_detailed_report(df_group):
    """Create detailed report with all original columns, formatting Order Date as long date with hour"""
    df = df_group.copy()
    
    if 'Order Date' in df.columns and pd.api.types.is_datetime64_any_dtype(df['Order Date']):
        df['Order Date'] = df['Order Date'].dt.strftime('%Y-%m-%d, %H:%M')
    
    if 'Product/Barcode' in df.columns:
        df['Product/Barcode'] = df['Product/Barcode'].astype(str)
    
    return df

def create_grouped_detailed_report(df_group, organize_by_brand=False):
    """Create detailed report grouped by brand with bold headers, sorted by date/time"""
    df = df_group.copy()
    
    # Remove Parent Brand column if it exists
    if 'Parent Brand' in df.columns:
        df = df.drop(columns=['Parent Brand'])
    
    # Ensure Order Date is datetime for proper sorting
    if 'Order Date' in df.columns:
        if not pd.api.types.is_datetime64_any_dtype(df['Order Date']):
            df['Order Date'] = pd.to_datetime(df['Order Date'])
    
    # Sort by date, then hour, then minute
    sort_columns = []
    if 'Order Date' in df.columns:
        sort_columns.append('Order Date')
    if 'Order Time' in df.columns:
        sort_columns.append('Order Time')
    
    if sort_columns:
        df = df.sort_values(by=sort_columns)
    
    # Format Order Date for display
    if 'Order Date' in df.columns:
        df['Order Date'] = df['Order Date'].dt.strftime('%Y-%m-%d, %H:%M')
    
    if 'Product/Barcode' in df.columns:
        df['Product/Barcode'] = df['Product/Barcode'].astype(str)
    
    # If organizing by brand, create grouped data
    if organize_by_brand and 'Brand' in df.columns:
        grouped_data = []
        brands = sorted(df['Brand'].dropna().unique())
        
        for brand in brands:
            brand_data = df[df['Brand'] == brand].copy()
            
            # Calculate totals for the brand
            total_quantity = brand_data['Quantity'].sum() if 'Quantity' in brand_data.columns else 0
            total_value = brand_data['Tax Incl.'].sum() if 'Tax Incl.' in brand_data.columns else 0
            
            # Format total value as Indonesian Rupiah
            if total_value == 0:
                formatted_total_value = "Rp 0.00"
            else:
                formatted_total_value = f"Rp {total_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
            # Add brand header row with totals
            brand_header = {}
            for col in brand_data.columns:
                if col == 'Brand':
                    brand_header[col] = f"{brand} - Total"
                elif col == 'Quantity':
                    brand_header[col] = total_quantity
                elif col == 'Tax Incl.':
                    brand_header[col] = formatted_total_value
                else:
                    brand_header[col] = ""
            grouped_data.append(brand_header)
            
            # Add brand data rows
            for _, row in brand_data.iterrows():
                grouped_data.append(row.to_dict())
        
        return pd.DataFrame(grouped_data)
    else:
        return df

def apply_excel_formatting(worksheet, data_type="pivot"):
    """Apply consistent formatting to Excel worksheet"""
    # Apply yellow fill to headers
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    bold_font = Font(bold=True)
    
    header_row = list(worksheet.iter_rows(min_row=1, max_row=1))[0]
    
    # Apply formatting to each header cell
    for cell in header_row:
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')
    
    # Apply currency formatting to monetary columns
    total_col_indices = []
    for idx, cell in enumerate(header_row, 1):
        if cell.value:
            col_name = str(cell.value)
            # Check for various monetary column patterns
            if (col_name.startswith('Total_') or 
                col_name.startswith('Total') or 
                'Tax Incl' in col_name or 
                'Revenue' in col_name or
                'Amount' in col_name or
                'Price' in col_name or
                '_Total_' in col_name):
                total_col_indices.append(idx)
    
    if total_col_indices:
        # Pre-format values as text with Rupiah symbol
        max_row = worksheet.max_row
        for row_idx in range(2, max_row + 1):
            for col_idx in total_col_indices:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    try:
                        if isinstance(cell.value, str):
                            cell.value = float(cell.value)
                        
                        # Format as Indonesian Rupiah text
                        if cell.value == 0:
                            formatted_value = "Rp 0.00"
                        else:
                            formatted_value = f"Rp {cell.value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        
                        cell.value = formatted_value
                        cell.number_format = '@'  # Text format
                    except:
                        pass
    
    # Format barcode columns as text
    barcode_col_idx = None
    for idx, cell in enumerate(header_row):
        if cell.value in ['Barcode', 'Product/Barcode', 'Product Barcode']:
            barcode_col_idx = idx
            break
    
    if barcode_col_idx is not None:
        for row in worksheet.iter_rows(min_row=2):
            cell = row[barcode_col_idx]
            if cell.value is not None:
                cell.value = str(cell.value)
                cell.number_format = '@'
    
    # Make brand total sellout rows bold
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        first_cell = row[0]
        if first_cell.value and (
            str(first_cell.value).endswith('Total Sellout') or
            str(first_cell.value) == 'Total Sellout'
        ):
            bold_font = Font(bold=True)
            for cell in row:
                cell.font = bold_font
    
    # Make brand header rows bold for detailed reports
    if data_type == "detailed":
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            brand_cell = None
            # Find the Brand column
            for cell in row:
                header_cell = worksheet.cell(row=1, column=cell.column)
                if header_cell.value == 'Brand':
                    brand_cell = cell
                    break
            
            if brand_cell and brand_cell.value and (
                str(brand_cell.value).endswith(' - Total') or
                str(brand_cell.value).endswith('Total Sellout')
            ):
                bold_font = Font(bold=True)
                # Apply light gray background to brand header
                header_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
                for cell in row:
                    if cell.value or str(brand_cell.value).endswith(' - Total'):  # Format all cells in total row
                        cell.font = bold_font
                        cell.fill = header_fill
    
    # Adjust column widths for better visibility
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Apply borders to all cells
    thin_border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border

def create_workbook_for_parent_brand(pivot_df, detailed_df, parent_brand_name, date_str=None, is_brand_organized=False, df_group=None, organize_by_brand=False, separate_by_date=False):
    """Create Excel workbook with 2 sheets: Pivoted and Detailed Report"""
    wb = Workbook()
    wb.remove(wb.active)
    
    if separate_by_date and df_group is not None:
        df_group['Order Date Day'] = pd.to_datetime(df_group['Order Date']).dt.normalize()
        unique_dates = sorted(df_group['Order Date Day'].unique())
        
        actual_parent_brand = parent_brand_name.split('_')[0] if '_' in parent_brand_name else parent_brand_name
        is_non_paragon_hebe = (actual_parent_brand not in ['Paragon', 'Hebe'])
        
        for date_val in unique_dates:
            date_df = df_group[df_group['Order Date Day'] == date_val].copy()
            
            if not date_df.empty:
                date_pivot_df = create_pivot_by_barcode(date_df)
                date_detailed_df = create_grouped_detailed_report(date_df, organize_by_brand=organize_by_brand)
                date_str_sheet = pd.to_datetime(date_val).strftime('%Y-%m-%d')
                
                ws_pivot = wb.create_sheet(f"Pivoted_{date_str_sheet}")
                
                if organize_by_brand and is_non_paragon_hebe and 'Brand' in date_df.columns:
                    brands = sorted(date_df['Brand'].dropna().unique())
                    
                    header_row_data = {}
                    for col in date_pivot_df.columns:
                        header_row_data[col] = col
                    ws_pivot.append([header_row_data.get(col, "") for col in date_pivot_df.columns])
                    
                    current_row = 2
                    
                    for brand in brands:
                        brand_products = date_df[date_df['Brand'] == brand]
                        
                        if not brand_products.empty:
                            brand_barcodes = set(brand_products['Product/Barcode'].astype(str))
                            brand_pivot = date_pivot_df[date_pivot_df['Barcode'].astype(str).isin(brand_barcodes)]
                            
                            total_cols = [col for col in date_pivot_df.columns if col.startswith('Total_')]
                            brand_total_sellout = brand_pivot[total_cols].sum().sum() if not brand_pivot.empty and total_cols else 0
                            
                            qty_cols = [col for col in date_pivot_df.columns if col.startswith('Jumlah_')]
                            brand_total_qty = brand_pivot[qty_cols].sum().sum() if not brand_pivot.empty and qty_cols else 0
                            
                            brand_total_row_data = {}
                            for col in date_pivot_df.columns:
                                if col in ['Barcode', 'Produk']:
                                    brand_total_row_data[col] = f"{brand} Total Sellout" if col == 'Barcode' else ""
                                elif col.startswith('Total_'):
                                    brand_total_row_data[col] = brand_total_sellout
                                elif col.startswith('Jumlah_'):
                                    brand_total_row_data[col] = brand_total_qty
                                else:
                                    brand_total_row_data[col] = ""
                            
                            ws_pivot.append([brand_total_row_data.get(col, "") for col in date_pivot_df.columns])
                            current_row += 1
                            
                            if not brand_pivot.empty:
                                for _, row in brand_pivot.iterrows():
                                    row_data = []
                                    for col in date_pivot_df.columns:
                                        row_data.append(row.get(col, ""))
                                    ws_pivot.append(row_data)
                                    current_row += 1
                else:
                    for r in dataframe_to_rows(date_pivot_df, index=False, header=True):
                        ws_pivot.append(r)
                    
                    total_row_data = {}
                    for col in date_pivot_df.columns:
                        if col in ['Barcode', 'Produk']:
                            total_row_data[col] = "Total Sellout" if col == 'Barcode' else ""
                        elif col.startswith('Total_'):
                            total_row_data[col] = date_pivot_df[col].sum()
                        elif col.startswith('Jumlah_'):
                            total_row_data[col] = date_pivot_df[col].sum()
                        else:
                            total_row_data[col] = ""
                    
                    total_row_df = pd.DataFrame([total_row_data])
                    total_row_df = total_row_df.reindex(columns=date_pivot_df.columns, fill_value="")
                    
                    ws_pivot.insert_rows(2)
                    for col_idx, col_name in enumerate(date_pivot_df.columns, 1):
                        cell = ws_pivot.cell(row=2, column=col_idx)
                        value = total_row_df.iloc[0][col_name]
                        cell.value = value
                
                # Apply formatting to pivot sheet
                apply_excel_formatting(ws_pivot, "pivot")
                
                ws_detailed = wb.create_sheet(f"Detailed Report_{date_str_sheet}")
                date_detailed_df = create_grouped_detailed_report(date_df, organize_by_brand=True)
                for r in dataframe_to_rows(date_detailed_df, index=False, header=True):
                    ws_detailed.append(r)
                
                # Apply formatting to detailed report sheet
                apply_excel_formatting(ws_detailed, "detailed")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    else:
        ws_pivot = wb.create_sheet("Pivoted")
        
        if not is_brand_organized:
            has_brand_prefixes = any('_' in col and col not in ['Barcode', 'Produk'] and 
                                     not col.startswith('Jumlah_') and not col.startswith('Total_') 
                                     for col in pivot_df.columns)
            is_brand_organized = has_brand_prefixes
        
        if organize_by_brand and df_group is not None and 'Brand' in df_group.columns:
            brands = sorted(df_group['Brand'].dropna().unique())
            
            header_row_data = {}
            for col in pivot_df.columns:
                header_row_data[col] = col
            ws_pivot.append([header_row_data.get(col, "") for col in pivot_df.columns])
            
            current_row = 2
            
            for brand in brands:
                brand_products = df_group[df_group['Brand'] == brand]
                
                if not brand_products.empty:
                    brand_barcodes = set(brand_products['Product/Barcode'].astype(str))
                    brand_pivot = pivot_df[pivot_df['Barcode'].astype(str).isin(brand_barcodes)]
                    
                    total_cols = [col for col in pivot_df.columns if col.startswith('Total_')]
                    brand_total_sellout = brand_pivot[total_cols].sum().sum() if not brand_pivot.empty and total_cols else 0
                    
                    qty_cols = [col for col in pivot_df.columns if col.startswith('Jumlah_')]
                    brand_total_qty = brand_pivot[qty_cols].sum().sum() if not brand_pivot.empty and qty_cols else 0
                    
                    brand_total_row_data = {}
                    for col in pivot_df.columns:
                        if col in ['Barcode', 'Produk']:
                            brand_total_row_data[col] = f"{brand} Total Sellout" if col == 'Barcode' else ""
                        elif col.startswith('Total_'):
                            brand_total_row_data[col] = brand_total_sellout
                        elif col.startswith('Jumlah_'):
                            brand_total_row_data[col] = brand_total_qty
                        else:
                            brand_total_row_data[col] = ""
                    
                    ws_pivot.append([brand_total_row_data.get(col, "") for col in pivot_df.columns])
                    current_row += 1
                    
                    if not brand_pivot.empty:
                        for _, row in brand_pivot.iterrows():
                            row_data = []
                            for col in pivot_df.columns:
                                row_data.append(row.get(col, ""))
                            ws_pivot.append(row_data)
                            current_row += 1
        else:
            for r in dataframe_to_rows(pivot_df, index=False, header=True):
                ws_pivot.append(r)
        
        if not organize_by_brand and not is_brand_organized:
            current_row = 2
            
            if df_group is not None and 'Brand' in df_group.columns:
                brands = sorted(df_group['Brand'].dropna().unique())
                
                for brand in brands:
                    brand_products = df_group[df_group['Brand'] == brand]
                    
                    if not brand_products.empty:
                        brand_barcodes = set(brand_products['Product/Barcode'].astype(str))
                        brand_pivot = pivot_df[pivot_df['Barcode'].astype(str).isin(brand_barcodes)]
                        
                        total_cols = [col for col in pivot_df.columns if col.startswith('Total_')]
                        brand_total_sellout = brand_pivot[total_cols].sum().sum() if not brand_pivot.empty and total_cols else 0
                        
                        qty_cols = [col for col in pivot_df.columns if col.startswith('Jumlah_')]
                        brand_total_qty = brand_pivot[qty_cols].sum().sum() if not brand_pivot.empty and qty_cols else 0
                        
                        brand_total_row_data = {}
                        for col in pivot_df.columns:
                            if col in ['Barcode', 'Produk']:
                                brand_total_row_data[col] = f"{brand} Total Sellout" if col == 'Barcode' else ""
                            elif col.startswith('Total_'):
                                brand_total_row_data[col] = brand_total_sellout
                            elif col.startswith('Jumlah_'):
                                brand_total_row_data[col] = brand_total_qty
                            else:
                                brand_total_row_data[col] = ""
                        
                        ws_pivot.insert_rows(current_row)
                        for col_idx, col_name in enumerate(pivot_df.columns, 1):
                            cell = ws_pivot.cell(row=current_row, column=col_idx)
                            value = brand_total_row_data.get(col_name, "")
                            cell.value = value
                        
                        current_row += 1
            
            total_row_data = {}
            
            for col in ['Barcode', 'Produk']:
                if col in pivot_df.columns:
                    total_row_data[col] = "Total Sellout" if col == 'Barcode' else ""
            
            total_cols = [col for col in pivot_df.columns if col.startswith('Total_')]
            for col in total_cols:
                total_row_data[col] = pivot_df[col].sum()
            
            qty_cols = [col for col in pivot_df.columns if col.startswith('Jumlah_')]
            for col in qty_cols:
                total_row_data[col] = pivot_df[col].sum()
            
            total_row_df = pd.DataFrame([total_row_data])
            total_row_df = total_row_df.reindex(columns=pivot_df.columns, fill_value="")
            
            ws_pivot.insert_rows(current_row)
            
            for col_idx, col_name in enumerate(pivot_df.columns, 1):
                cell = ws_pivot.cell(row=current_row, column=col_idx)
                value = total_row_df.iloc[0][col_name]
                cell.value = value
        
        # Apply formatting to pivot sheet
        apply_excel_formatting(ws_pivot, "pivot")
        
        ws_detailed = wb.create_sheet("Detailed Report")
        detailed_df = create_grouped_detailed_report(df_group, organize_by_brand=organize_by_brand)
        for r in dataframe_to_rows(detailed_df, index=False, header=True):
            ws_detailed.append(r)
        
        # Apply formatting to detailed report sheet
        apply_excel_formatting(ws_detailed, "detailed")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

def create_zip_file(workbooks_dict):
    """Create zip file containing all workbooks"""
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for workbook_key, workbook_bytes in workbooks_dict.items():
            sanitized_name = sanitize_filename(workbook_key)
            filename = f"{sanitized_name}.xlsx"
            zip_file.writestr(filename, workbook_bytes.getvalue())
    
    zip_buffer.seek(0)
    return zip_buffer
