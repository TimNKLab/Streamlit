import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import zipfile
from datetime import datetime
import re

# Configure page
st.set_page_config(
    page_title="NK Lab",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Initialize session state for authentication
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False

# Simple password authentication
PASSWORD = "admin123"

def login():
    """Display login page"""
    st.title("🔐 Authentication Required")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.markdown("### Masukkan Password untuk melanjutkan")
        password = st.text_input("Password", type="password", key="password_input")
        
        if st.button("Login", type="primary", use_container_width=True):
            if password == PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("❌ Password salah. Silakan coba lagi.")

def dashboard_page():
    """Dashboard page content"""
    st.title("📊 Dashboard")
    st.markdown("### NK Dashboard v0.1.3")
    
    st.info("This page will display an overview of key business metrics and KPIs.")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Sales", "---", "---")
    
    with col2:
        st.metric("Active Users", "---", "---")
    
    with col3:
        st.metric("Inventory Value", "---", "---")
    
    with col4:
        st.metric("Orders", "---", "---")
    
    st.markdown("---")
    st.subheader("📈 Charts and Visualizations")
    st.text("Interactive charts and graphs will be displayed here.")
    
    st.markdown("---")
    st.subheader("📋 Recent Activity")
    st.text("Recent transactions and updates will be shown in this section.")

def sanitize_filename(name):
    """Sanitize filename by removing invalid characters"""
    invalid_chars = r'[<>:"/\\|?*]'
    sanitized = re.sub(invalid_chars, '_', str(name))
    sanitized = sanitized.strip(' .')
    return sanitized if sanitized else 'Unknown'

def sort_sales_data(df):
    """Sort data by Parent Brand (alphabetically) then Order Date (earliest date, then earliest hour)"""
    df = df.copy()
    
    if not pd.api.types.is_datetime64_any_dtype(df['Order Date']):
        df['Order Date'] = pd.to_datetime(df['Order Date'], errors='coerce')
    
    df['sort_key_parent_brand'] = df['Parent Brand'].fillna(df['Brand'])
    df['sort_date'] = df['Order Date'].dt.date
    df['sort_hour'] = df['Order Date'].dt.hour
    
    df_sorted = df.sort_values(
        by=['sort_key_parent_brand', 'sort_date', 'sort_hour', 'Order Date'],
        ascending=[True, True, True, True],
        na_position='last'
    )
    
    df_sorted = df_sorted.drop(columns=['sort_key_parent_brand', 'sort_date', 'sort_hour'])
    
    return df_sorted.reset_index(drop=True)

def group_by_parent_brand(df):
    """Group data by Parent Brand (use Brand if Parent Brand is None).
    For Paragon and Hebe, also split by Brand."""
    groups = {}
    split_by_brand_parents = ['Paragon', 'Hebe']
    
    for idx, row in df.iterrows():
        parent_brand = row['Parent Brand'] if pd.notna(row['Parent Brand']) else row['Brand']
        brand = row['Brand'] if pd.notna(row['Brand']) else 'Unknown'
        
        if parent_brand in split_by_brand_parents:
            group_key = f"{parent_brand}_{brand}"
        else:
            group_key = parent_brand
        
        if group_key not in groups:
            groups[group_key] = []
        groups[group_key].append(idx)
    
    grouped_dfs = {}
    for key, indices in groups.items():
        grouped_dfs[key] = df.loc[indices].reset_index(drop=True)
    
    return grouped_dfs

def create_pivot_by_barcode(df_group):
    """Create pivot table with barcode as index, dates as columns, showing Quantity and Tax Incl."""
    df = df_group.copy()
    
    df['Product/Barcode'] = df['Product/Barcode'].astype(str)
    
    if not pd.api.types.is_datetime64_any_dtype(df['Order Date']):
        df['Order Date'] = pd.to_datetime(df['Order Date'])
    
    df['Order Date Day'] = pd.to_datetime(df['Order Date']).dt.normalize()
    
    pivot_qty = pd.pivot_table(
        df,
        index=['Product/Barcode', 'Product'],
        columns='Order Date Day',
        values='Quantity',
        aggfunc='sum',
        fill_value=0
    )
    
    pivot_revenue = pd.pivot_table(
        df,
        index=['Product/Barcode', 'Product'],
        columns='Order Date Day',
        values='Tax Incl.',
        aggfunc='sum',
        fill_value=0
    )
    
    def format_col_name(prefix, col_val):
        if pd.notna(col_val) and isinstance(col_val, pd.Timestamp):
            return f"{prefix}_{col_val.strftime('%Y-%m-%d')}"
        else:
            return f"{prefix}_{str(col_val)}"
    
    pivot_qty.columns = [format_col_name("Jumlah", col) for col in pivot_qty.columns]
    pivot_revenue.columns = [format_col_name("Total", col) for col in pivot_revenue.columns]
    
    pivot = pd.merge(
        pivot_qty.reset_index(),
        pivot_revenue.reset_index(),
        on=['Product/Barcode', 'Product'],
        how='outer'
    )
    
    pivot = pivot.rename(columns={'Product/Barcode': 'Barcode', 'Product': 'Produk'})
    
    base_cols = ['Barcode', 'Produk']
    qty_cols = [col for col in pivot.columns if col.startswith('Jumlah_')]
    revenue_cols = [col for col in pivot.columns if col.startswith('Total_')]
    
    def extract_date(col_name):
        try:
            date_str = col_name.split('_', 1)[1]
            return pd.to_datetime(date_str)
        except:
            return pd.Timestamp.min
    
    qty_cols_sorted = sorted(qty_cols, key=extract_date)
    revenue_cols_sorted = sorted(revenue_cols, key=extract_date)
    
    pivot = pivot[base_cols + qty_cols_sorted + revenue_cols_sorted]
    
    return pivot

def create_detailed_report(df_group):
    """Create detailed report with all original columns, formatting Order Date as long date with hour"""
    df = df_group.copy()
    
    if 'Order Date' in df.columns and pd.api.types.is_datetime64_any_dtype(df['Order Date']):
        df['Order Date'] = df['Order Date'].dt.strftime('%Y-%m-%d, %H:%M')
    
    if 'Product/Barcode' in df.columns:
        df['Product/Barcode'] = df['Product/Barcode'].astype(str)
    
    return df

def create_grouped_detailed_report(df_group, organize_by_brand=False):
    """Create detailed report grouped by brand with bold headers, sorted by date/time"""
    df = df_group.copy()
    
    # Remove Parent Brand column if it exists
    if 'Parent Brand' in df.columns:
        df = df.drop(columns=['Parent Brand'])
    
    # Ensure Order Date is datetime for proper sorting
    if 'Order Date' in df.columns:
        if not pd.api.types.is_datetime64_any_dtype(df['Order Date']):
            df['Order Date'] = pd.to_datetime(df['Order Date'])
    
    # Sort by date, then hour, then minute
    sort_columns = []
    if 'Order Date' in df.columns:
        sort_columns.append('Order Date')
    if 'Order Time' in df.columns:
        sort_columns.append('Order Time')
    
    if sort_columns:
        df = df.sort_values(by=sort_columns)
    
    # Format Order Date for display
    if 'Order Date' in df.columns:
        df['Order Date'] = df['Order Date'].dt.strftime('%Y-%m-%d, %H:%M')
    
    if 'Product/Barcode' in df.columns:
        df['Product/Barcode'] = df['Product/Barcode'].astype(str)
    
    # If organizing by brand, create grouped data
    if organize_by_brand and 'Brand' in df.columns:
        grouped_data = []
        brands = sorted(df['Brand'].dropna().unique())
        
        for brand in brands:
            brand_data = df[df['Brand'] == brand].copy()
            
            # Calculate totals for the brand
            total_quantity = brand_data['Quantity'].sum() if 'Quantity' in brand_data.columns else 0
            total_value = brand_data['Tax Incl.'].sum() if 'Tax Incl.' in brand_data.columns else 0
            
            # Format total value as Indonesian Rupiah
            if total_value == 0:
                formatted_total_value = "Rp 0.00"
            else:
                formatted_total_value = f"Rp {total_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
            # Add brand header row with totals
            brand_header = {}
            for col in brand_data.columns:
                if col == 'Brand':
                    brand_header[col] = f"{brand} - Total"
                elif col == 'Quantity':
                    brand_header[col] = total_quantity
                elif col == 'Tax Incl.':
                    brand_header[col] = formatted_total_value
                else:
                    brand_header[col] = ""
            grouped_data.append(brand_header)
            
            # Add brand data rows
            for _, row in brand_data.iterrows():
                grouped_data.append(row.to_dict())
        
        return pd.DataFrame(grouped_data)
    else:
        return df

def apply_excel_formatting(worksheet, data_type="pivot"):
    """Apply consistent formatting to Excel worksheet"""
    # Apply yellow fill to headers
    yellow_fill = PatternFill(start_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    
    header_row = list(worksheet.iter_rows(min_row=1, max_row=1))[0]
    
    # Apply formatting to each header cell
    for cell in header_row:
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')
    
    # Apply currency formatting to monetary columns
    total_col_indices = []
    for idx, cell in enumerate(header_row, 1):
        if cell.value:
            col_name = str(cell.value)
            # Check for various monetary column patterns
            if (col_name.startswith('Total_') or 
                col_name.startswith('Total') or 
                'Tax Incl' in col_name or 
                'Revenue' in col_name or
                'Amount' in col_name or
                'Price' in col_name or
                '_Total_' in col_name):
                total_col_indices.append(idx)
    
    if total_col_indices:
        # Pre-format values as text with Rupiah symbol
        max_row = worksheet.max_row
        for row_idx in range(2, max_row + 1):
            for col_idx in total_col_indices:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    try:
                        if isinstance(cell.value, str):
                            cell.value = float(cell.value)
                        
                        # Format as Indonesian Rupiah text
                        if cell.value == 0:
                            formatted_value = "Rp 0.00"
                        else:
                            formatted_value = f"Rp {cell.value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        
                        cell.value = formatted_value
                        cell.number_format = '@'  # Text format
                    except:
                        pass
    
    # Format barcode columns as text
    barcode_col_idx = None
    for idx, cell in enumerate(header_row):
        if cell.value in ['Barcode', 'Product/Barcode', 'Product Barcode']:
            barcode_col_idx = idx
            break
    
    if barcode_col_idx is not None:
        for row in worksheet.iter_rows(min_row=2):
            cell = row[barcode_col_idx]
            if cell.value is not None:
                cell.value = str(cell.value)
                cell.number_format = '@'
    
    # Make brand total sellout rows bold
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        first_cell = row[0]
        if first_cell.value and (
            str(first_cell.value).endswith('Total Sellout') or
            str(first_cell.value) == 'Total Sellout'
        ):
            bold_font = Font(bold=True)
            for cell in row:
                cell.font = bold_font
    
    # Make brand header rows bold for detailed reports
    if data_type == "detailed":
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            brand_cell = None
            # Find the Brand column
            for cell in row:
                header_cell = worksheet.cell(row=1, column=cell.column)
                if header_cell.value == 'Brand':
                    brand_cell = cell
                    break
            
            if brand_cell and brand_cell.value and (
                str(brand_cell.value).endswith(' - Total') or
                str(brand_cell.value).endswith('Total Sellout')
            ):
                bold_font = Font(bold=True)
                # Apply light gray background to brand header
                header_fill = PatternFill(start_color="F2F2F2", fill_type="solid")
                for cell in row:
                    if cell.value or str(brand_cell.value).endswith(' - Total'):  # Format all cells in total row
                        cell.font = bold_font
                        cell.fill = header_fill
    
    # Adjust column widths for better visibility
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Apply borders to all cells
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border

def create_workbook_for_parent_brand(pivot_df, detailed_df, parent_brand_name, date_str=None, is_brand_organized=False, df_group=None, organize_by_brand=False, separate_by_date=False):
    """Create Excel workbook with 2 sheets: Pivoted and Detailed Report"""
    wb = Workbook()
    wb.remove(wb.active)
    
    if separate_by_date and df_group is not None:
        df_group['Order Date Day'] = pd.to_datetime(df_group['Order Date']).dt.normalize()
        unique_dates = sorted(df_group['Order Date Day'].unique())
        
        actual_parent_brand = parent_brand_name.split('_')[0] if '_' in parent_brand_name else parent_brand_name
        is_non_paragon_hebe = (actual_parent_brand not in ['Paragon', 'Hebe'])
        
        for date_val in unique_dates:
            date_df = df_group[df_group['Order Date Day'] == date_val].copy()
            
            if not date_df.empty:
                date_pivot_df = create_pivot_by_barcode(date_df)
                date_detailed_df = create_grouped_detailed_report(date_df, organize_by_brand=organize_by_brand)
                date_str_sheet = pd.to_datetime(date_val).strftime('%Y-%m-%d')
                
                ws_pivot = wb.create_sheet(f"Pivoted_{date_str_sheet}")
                
                if organize_by_brand and is_non_paragon_hebe and 'Brand' in date_df.columns:
                    brands = sorted(date_df['Brand'].dropna().unique())
                    
                    header_row_data = {}
                    for col in date_pivot_df.columns:
                        header_row_data[col] = col
                    ws_pivot.append([header_row_data.get(col, "") for col in date_pivot_df.columns])
                    
                    current_row = 2
                    
                    for brand in brands:
                        brand_products = date_df[date_df['Brand'] == brand]
                        
                        if not brand_products.empty:
                            brand_barcodes = set(brand_products['Product/Barcode'].astype(str))
                            brand_pivot = date_pivot_df[date_pivot_df['Barcode'].astype(str).isin(brand_barcodes)]
                            
                            total_cols = [col for col in date_pivot_df.columns if col.startswith('Total_')]
                            brand_total_sellout = brand_pivot[total_cols].sum().sum() if not brand_pivot.empty and total_cols else 0
                            
                            qty_cols = [col for col in date_pivot_df.columns if col.startswith('Jumlah_')]
                            brand_total_qty = brand_pivot[qty_cols].sum().sum() if not brand_pivot.empty and qty_cols else 0
                            
                            brand_total_row_data = {}
                            for col in date_pivot_df.columns:
                                if col in ['Barcode', 'Produk']:
                                    brand_total_row_data[col] = f"{brand} Total Sellout" if col == 'Barcode' else ""
                                elif col.startswith('Total_'):
                                    brand_total_row_data[col] = brand_total_sellout
                                elif col.startswith('Jumlah_'):
                                    brand_total_row_data[col] = brand_total_qty
                                else:
                                    brand_total_row_data[col] = ""
                            
                            ws_pivot.append([brand_total_row_data.get(col, "") for col in date_pivot_df.columns])
                            current_row += 1
                            
                            if not brand_pivot.empty:
                                for _, row in brand_pivot.iterrows():
                                    row_data = []
                                    for col in date_pivot_df.columns:
                                        row_data.append(row.get(col, ""))
                                    ws_pivot.append(row_data)
                                    current_row += 1
                else:
                    for r in dataframe_to_rows(date_pivot_df, index=False, header=True):
                        ws_pivot.append(r)
                    
                    total_row_data = {}
                    for col in date_pivot_df.columns:
                        if col in ['Barcode', 'Produk']:
                            total_row_data[col] = "Total Sellout" if col == 'Barcode' else ""
                        elif col.startswith('Total_'):
                            total_row_data[col] = date_pivot_df[col].sum()
                        elif col.startswith('Jumlah_'):
                            total_row_data[col] = date_pivot_df[col].sum()
                        else:
                            total_row_data[col] = ""
                    
                    total_row_df = pd.DataFrame([total_row_data])
                    total_row_df = total_row_df.reindex(columns=date_pivot_df.columns, fill_value="")
                    
                    ws_pivot.insert_rows(2)
                    for col_idx, col_name in enumerate(date_pivot_df.columns, 1):
                        cell = ws_pivot.cell(row=2, column=col_idx)
                        value = total_row_df.iloc[0][col_name]
                        cell.value = value
                
                # Apply formatting to pivot sheet
                apply_excel_formatting(ws_pivot, "pivot")
                
                ws_detailed = wb.create_sheet(f"Detailed Report_{date_str_sheet}")
                date_detailed_df = create_grouped_detailed_report(date_df, organize_by_brand=True)
                for r in dataframe_to_rows(date_detailed_df, index=False, header=True):
                    ws_detailed.append(r)
                
                # Apply formatting to detailed report sheet
                apply_excel_formatting(ws_detailed, "detailed")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    else:
        ws_pivot = wb.create_sheet("Pivoted")
        
        if not is_brand_organized:
            has_brand_prefixes = any('_' in col and col not in ['Barcode', 'Produk'] and 
                                     not col.startswith('Jumlah_') and not col.startswith('Total_') 
                                     for col in pivot_df.columns)
            is_brand_organized = has_brand_prefixes
        
        if organize_by_brand and df_group is not None and 'Brand' in df_group.columns:
            brands = sorted(df_group['Brand'].dropna().unique())
            
            header_row_data = {}
            for col in pivot_df.columns:
                header_row_data[col] = col
            ws_pivot.append([header_row_data.get(col, "") for col in pivot_df.columns])
            
            current_row = 2
            
            for brand in brands:
                brand_products = df_group[df_group['Brand'] == brand]
                
                if not brand_products.empty:
                    brand_barcodes = set(brand_products['Product/Barcode'].astype(str))
                    brand_pivot = pivot_df[pivot_df['Barcode'].astype(str).isin(brand_barcodes)]
                    
                    total_cols = [col for col in pivot_df.columns if col.startswith('Total_')]
                    brand_total_sellout = brand_pivot[total_cols].sum().sum() if not brand_pivot.empty and total_cols else 0
                    
                    qty_cols = [col for col in pivot_df.columns if col.startswith('Jumlah_')]
                    brand_total_qty = brand_pivot[qty_cols].sum().sum() if not brand_pivot.empty and qty_cols else 0
                    
                    brand_total_row_data = {}
                    for col in pivot_df.columns:
                        if col in ['Barcode', 'Produk']:
                            brand_total_row_data[col] = f"{brand} Total Sellout" if col == 'Barcode' else ""
                        elif col.startswith('Total_'):
                            brand_total_row_data[col] = brand_total_sellout
                        elif col.startswith('Jumlah_'):
                            brand_total_row_data[col] = brand_total_qty
                        else:
                            brand_total_row_data[col] = ""
                    
                    ws_pivot.append([brand_total_row_data.get(col, "") for col in pivot_df.columns])
                    current_row += 1
                    
                    if not brand_pivot.empty:
                        for _, row in brand_pivot.iterrows():
                            row_data = []
                            for col in pivot_df.columns:
                                row_data.append(row.get(col, ""))
                            ws_pivot.append(row_data)
                            current_row += 1
        else:
            for r in dataframe_to_rows(pivot_df, index=False, header=True):
                ws_pivot.append(r)
        
        if not organize_by_brand and not is_brand_organized:
            current_row = 2
            
            if df_group is not None and 'Brand' in df_group.columns:
                brands = sorted(df_group['Brand'].dropna().unique())
                
                for brand in brands:
                    brand_products = df_group[df_group['Brand'] == brand]
                    
                    if not brand_products.empty:
                        brand_barcodes = set(brand_products['Product/Barcode'].astype(str))
                        brand_pivot = pivot_df[pivot_df['Barcode'].astype(str).isin(brand_barcodes)]
                        
                        total_cols = [col for col in pivot_df.columns if col.startswith('Total_')]
                        brand_total_sellout = brand_pivot[total_cols].sum().sum() if not brand_pivot.empty and total_cols else 0
                        
                        qty_cols = [col for col in pivot_df.columns if col.startswith('Jumlah_')]
                        brand_total_qty = brand_pivot[qty_cols].sum().sum() if not brand_pivot.empty and qty_cols else 0
                        
                        brand_total_row_data = {}
                        for col in pivot_df.columns:
                            if col in ['Barcode', 'Produk']:
                                brand_total_row_data[col] = f"{brand} Total Sellout" if col == 'Barcode' else ""
                            elif col.startswith('Total_'):
                                brand_total_row_data[col] = brand_total_sellout
                            elif col.startswith('Jumlah_'):
                                brand_total_row_data[col] = brand_total_qty
                            else:
                                brand_total_row_data[col] = ""
                        
                        ws_pivot.insert_rows(current_row)
                        for col_idx, col_name in enumerate(pivot_df.columns, 1):
                            cell = ws_pivot.cell(row=current_row, column=col_idx)
                            value = brand_total_row_data.get(col_name, "")
                            cell.value = value
                        
                        current_row += 1
            
            total_row_data = {}
            
            for col in ['Barcode', 'Produk']:
                if col in pivot_df.columns:
                    total_row_data[col] = "Total Sellout" if col == 'Barcode' else ""
            
            total_cols = [col for col in pivot_df.columns if col.startswith('Total_')]
            for col in total_cols:
                total_row_data[col] = pivot_df[col].sum()
            
            qty_cols = [col for col in pivot_df.columns if col.startswith('Jumlah_')]
            for col in qty_cols:
                total_row_data[col] = pivot_df[col].sum()
            
            total_row_df = pd.DataFrame([total_row_data])
            total_row_df = total_row_df.reindex(columns=pivot_df.columns, fill_value="")
            
            ws_pivot.insert_rows(current_row)
            
            for col_idx, col_name in enumerate(pivot_df.columns, 1):
                cell = ws_pivot.cell(row=current_row, column=col_idx)
                value = total_row_df.iloc[0][col_name]
                cell.value = value
        
        # Apply formatting to pivot sheet
        apply_excel_formatting(ws_pivot, "pivot")
        
        ws_detailed = wb.create_sheet("Detailed Report")
        detailed_df = create_grouped_detailed_report(df_group, organize_by_brand=organize_by_brand)
        for r in dataframe_to_rows(detailed_df, index=False, header=True):
            ws_detailed.append(r)
        
        # Apply formatting to detailed report sheet
        apply_excel_formatting(ws_detailed, "detailed")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

def create_zip_file(workbooks_dict):
    """Create zip file containing all workbooks"""
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for workbook_key, workbook_bytes in workbooks_dict.items():
            sanitized_name = sanitize_filename(workbook_key)
            filename = f"{sanitized_name}.xlsx"
            zip_file.writestr(filename, workbook_bytes.getvalue())
    
    zip_buffer.seek(0)
    return zip_buffer

def process_sales_workbook(uploaded_file, separate_by_date=False):
    """Main processing function"""
    try:
        df = pd.read_excel(uploaded_file, dtype={'Product/Barcode': str})
        
        required_columns = ['Order Date', 'Product/Barcode', 'Product', 'Parent Brand', 'Brand', 'Quantity', 'Tax Incl.']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
        
        df['Product/Barcode'] = df['Product/Barcode'].astype(str)
        df['Order Date'] = pd.to_datetime(df['Order Date'], errors='coerce')
        df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0)
        df['Tax Incl.'] = pd.to_numeric(df['Tax Incl.'], errors='coerce').fillna(0)
        
        df = df.dropna(subset=['Order Date', 'Product/Barcode'])
        
        if df.empty:
            raise ValueError("No valid data found in the file")
        
        df_sorted = sort_sales_data(df)
        grouped_dfs = group_by_parent_brand(df_sorted)
        
        workbooks_dict = {}
        for parent_brand, df_group in grouped_dfs.items():
            earliest_date = df_group['Order Date'].min()
            latest_date = df_group['Order Date'].max()
            
            if earliest_date.date() == latest_date.date():
                date_str = earliest_date.strftime('%d%m%Y')
            else:
                date_str_start = earliest_date.strftime('%d%m%Y')
                date_str_end = latest_date.strftime('%d%m%Y')
                date_str = f"{date_str_start}_to_{date_str_end}"
            
            actual_parent_brand = parent_brand.split('_')[0] if '_' in parent_brand else parent_brand
            is_non_paragon_hebe = (actual_parent_brand not in ['Paragon', 'Hebe'])
            
            if separate_by_date:
                workbook_bytes = create_workbook_for_parent_brand(
                    None, None, parent_brand, date_str, 
                    is_brand_organized=False, 
                    df_group=df_group, 
                    organize_by_brand=is_non_paragon_hebe,
                    separate_by_date=True
                )
            else:
                pivot_df = create_pivot_by_barcode(df_group)
                detailed_df = create_grouped_detailed_report(df_group, organize_by_brand=is_non_paragon_hebe)
                
                workbook_bytes = create_workbook_for_parent_brand(
                    pivot_df, detailed_df, parent_brand, date_str, 
                    is_brand_organized=False, 
                    df_group=df_group, 
                    organize_by_brand=is_non_paragon_hebe,
                    separate_by_date=False
                )
            
            workbooks_dict[f"{parent_brand}_{date_str}"] = workbook_bytes
        
        zip_file = create_zip_file(workbooks_dict)
        
        unique_parent_brands = set()
        for group_key in grouped_dfs.keys():
            if '_' in group_key and any(pb in group_key for pb in ['Paragon', 'Hebe']):
                parent_brand = group_key.split('_')[0]
                unique_parent_brands.add(parent_brand)
            else:
                unique_parent_brands.add(group_key)
        
        summary = {
            'total_rows': len(df_sorted),
            'parent_brands_count': len(unique_parent_brands),
            'workbooks_count': len(workbooks_dict),
            'date_range': {
                'start': df_sorted['Order Date'].min().strftime('%Y-%m-%d'),
                'end': df_sorted['Order Date'].max().strftime('%Y-%m-%d')
            },
            'parent_brands': sorted(unique_parent_brands),
            'workbook_keys': sorted(workbooks_dict.keys())
        }
        
        return workbooks_dict, zip_file, summary, None
        
    except Exception as e:
        return None, None, None, str(e)

def ba_sales_report_page():
    """BA Sales Report page content"""
    st.title("💰 Laporan Sellout Beauty Advisor (BA)")
    st.markdown("### Pembuat Laporan Sellout untuk Beauty Advisor (BA) dengan data dari ERP.")
    
    if 'processed_workbooks' not in st.session_state:
        st.session_state.processed_workbooks = None
    if 'processed_zip' not in st.session_state:
        st.session_state.processed_zip = None
    if 'processing_summary' not in st.session_state:
        st.session_state.processing_summary = None
    if 'processing_error' not in st.session_state:
        st.session_state.processing_error = None
    if 'date_organization_option' not in st.session_state:
        st.session_state.date_organization_option = "Satukan tanggal"
    
    st.subheader("📅 Date Organization")
    date_option = st.radio(
        "Pilih cara pengorganisasian tanggal:",
        ["Satukan tanggal", "Pisahkan per tanggal"],
        index=0 if st.session_state.date_organization_option == "Satukan tanggal" else 1,
        help="Satukan tanggal: Semua tanggal dalam satu tabel pivot. Pisahkan per tanggal: Setiap tanggal memiliki sheet terpisah."
    )
    st.session_state.date_organization_option = date_option
    
    st.subheader("📤 Upload Excel File")
    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=['xlsx', 'xls'],
        help="Upload Excel file with sales data. Required columns: Order Date, Product/Barcode, Product, Parent Brand, Brand, Quantity, Tax Incl."
    )
    
    if uploaded_file is not None:
        col1, col2 = st.columns(2)
        with col1:
            st.info(f"📄 File: {uploaded_file.name}")
        with col2:
            file_size = uploaded_file.size / 1024
            st.info(f"📊 Size: {file_size:.2f} KB")
        
        if st.button("🔄 Process File", type="primary", use_container_width=True):
            with st.spinner("Processing file... This may take a moment."):
                separate_by_date = (st.session_state.date_organization_option == "Pisahkan per tanggal")
                workbooks_dict, zip_file, summary, error = process_sales_workbook(uploaded_file, separate_by_date=separate_by_date)
                
                if error:
                    st.session_state.processing_error = error
                    st.session_state.processed_workbooks = None
                    st.session_state.processed_zip = None
                    st.session_state.processing_summary = None
                else:
                    st.session_state.processed_workbooks = workbooks_dict
                    st.session_state.processed_zip = zip_file
                    st.session_state.processing_summary = summary
                    st.session_state.processing_error = None
                    st.success("✅ File processed successfully!")
                    st.rerun()
    
    if st.session_state.processing_error:
        st.error(f"❌ Error: {st.session_state.processing_error}")
    
    if st.session_state.processing_summary:
        st.markdown("---")
        st.subheader("📊 Processing Summary")
        
        summary = st.session_state.processing_summary
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            st.metric("Total Rows", f"{summary['total_rows']:,}")
        
        with col2:
            st.metric("Parent Brands", summary['parent_brands_count'])
        
        with col3:
            st.metric("Workbooks", summary.get('workbooks_count', summary['parent_brands_count']))
        
        with col4:
            st.metric("Start Date", summary['date_range']['start'])
        
        with col5:
            st.metric("End Date", summary['date_range']['end'])
        
        st.markdown("---")
        st.subheader("📥 Download Reports")
        
        if st.session_state.processed_zip:
            zip_bytes = st.session_state.processed_zip.getvalue()
            st.download_button(
                label="📦 Download All Workbooks (ZIP)",
                data=zip_bytes,
                file_name=f"Laporan Penjualan BA {datetime.now().strftime('%d%m%Y')}.zip",
                mime="application/zip",
                use_container_width=True,
                type="primary"
            )
        
        st.markdown("---")
        st.subheader("📄 Individual Brand Downloads")
        
        if st.session_state.processed_workbooks:
            cols = st.columns(3)
            for idx, (workbook_key, workbook_bytes) in enumerate(st.session_state.processed_workbooks.items()):
                col_idx = idx % 3
                with cols[col_idx]:
                    sanitized_name = sanitize_filename(workbook_key)
                    filename = f"{sanitized_name}.xlsx"
                    workbook_data = workbook_bytes.getvalue()
                    
                    display_label = workbook_key.replace('_', ' ')[:40]
                    
                    st.download_button(
                        label=f"📥 {display_label}",
                        data=workbook_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"download_{sanitized_name}_{idx}"
                    )
        
        st.markdown("---")
        if st.button("🔄 Process New File", use_container_width=True):
            st.session_state.processed_workbooks = None
            st.session_state.processed_zip = None
            st.session_state.processing_summary = None
            st.session_state.processing_error = None
            st.rerun()
    
    if not st.session_state.processing_summary and not uploaded_file:
        st.markdown("---")
        st.info("👆 Upload Excel laporan penjualan ke sini.")
        
        st.markdown("---")
        st.subheader("📋 Expected File Format")
        st.text("The Excel file should contain the following columns:")
        st.code("""
- Order Date (datetime)
- Product/Barcode (text)
- Product (text)
- Parent Brand (text, can be empty)
- Brand (text)
- Quantity (numeric)
- Tax Incl. (numeric)
        """)
        
        st.markdown("---")
        st.subheader("ℹ️ Cara Penggunaan")
        st.markdown("""
        1. **Upload**: Upload your Excel file with sales data
        2. **Process**: Click "Process File" to sort and group the data
        3. **Sorting**: Data is sorted by Parent Brand (alphabetically), then by Order Date (earliest first)
        4. **Grouping**: Data is split into separate workbooks by Parent Brand (uses Brand if Parent Brand is empty)
        5. **Reports**: Each workbook contains:
           - **Pivoted Sheet**: Pivot table with barcode as rows, dates as columns (grouped by day)
           - **Detailed Report Sheet**: All transactions for that Parent Brand
        6. **Download**: Download individual workbooks or all workbooks as a ZIP file
        """)

def stock_control_page():
    """Stock Control page content"""
    st.title("📦 Stock Control")
    st.markdown("### Inventory Management System")
    
    st.info("This page will provide tools for monitoring and managing inventory levels.")
    
    st.subheader("📊 Current Stock Levels")
    st.text("Real-time inventory status will be displayed here.")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown("#### DSI Report")
        st.text("DSI level will be flagged here.")
    
    with col2:
        st.markdown("#### Area Overstock Report")
        st.text("Untuk SKU dengan jumlah terlalu banyak terpajang di area.")
    
    with col3:
        st.markdown("#### Area Understock Report")
        st.text("Untuk SKU dengan jumlah terlalu sedikit terpajang di area.")

    with col4:
        st.markdown("#### ABC Analysis")
        st.text("Prioritas barang berdasarkan ABC Analysis.")
    
    st.markdown("---")
    st.subheader("🔍 Stock Search & Filter")
    st.text("Advanced search and filtering capabilities will be available here.")
    
    st.markdown("---")
    st.subheader("➕ Stock Operations")
    st.text("Stock adjustments, transfers, and other operations will be performed here.")

def dsi_report_page():
    """DSI Report page content"""
    st.title("📋 DSI Report")
    st.markdown("### Days Sales of Inventory Report")
    
    st.info("This page will display Days Sales of Inventory (DSI) analysis and metrics.")
    
    st.subheader("📊 DSI Overview")
    st.text("Key DSI metrics and indicators will be displayed here.")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### Current DSI")
        st.metric("Average DSI", "---", "---")
        st.text("Detailed DSI calculations will be shown here.")
    
    with col2:
        st.markdown("#### DSI Trends")
        st.text("Historical DSI trends and patterns will be visualized here.")
    
    st.markdown("---")
    st.subheader("📈 Analysis by Category")
    st.text("DSI breakdown by product category will be available here.")
    
    st.markdown("---")
    st.subheader("⚠️ Alerts & Recommendations")
    st.text("DSI-related alerts and optimization recommendations will be provided here.")

def main():
    """Main application"""
    if not st.session_state.authenticated:
        login()
        return
    
    st.sidebar.title("Navigation")
    
    if st.sidebar.button("🚪 Logout", use_container_width=True):
        st.session_state.authenticated = False
        st.rerun()
    
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Dashboard", "💰 BA Sales Report", "📦 Stock Control", "📋 DSI Report"])
    
    with tab1:
        dashboard_page()
    
    with tab2:
        ba_sales_report_page()
    
    with tab3:
        stock_control_page()
    
    with tab4:
        dsi_report_page()

if __name__ == "__main__":
    main()
    
    st.markdown(
        """
        <hr style="margin-top: 3em; margin-bottom: 0.5em;">
        <div style="text-align: center; color: gray; font-size: 0.95em;">
            Dibuat dengan ❤️, dari Tim Data NK.
        </div>
        """,
        unsafe_allow_html=True
    )